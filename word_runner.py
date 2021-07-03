from PyQt5.QtWidgets import QGridLayout, QLabel, QLineEdit, QMessageBox, QPushButton, QWidget, QApplication, QFileDialog
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QIcon

import os
import sys
import time
import openpyxl


MAX_VOCABULARY_LINE = 3647
VOCABULARY_PATH = "./toefl_vocabulary.xlsx"

TYPE_LEARN = "learned"
TYPE_REVIEW = "reviewed"

NO_JUMP = "not_jump"

def create_button(label, function, tip):
    btn = QPushButton(label)
    btn.clicked.connect(function)
    if tip:
        btn.setToolTip(tip)
    btn.resize(btn.sizeHint())

    return btn

def create_label(label):
    lab = QLabel()
    lab.setText(label)
    lab.adjustSize()

    return lab

def create_lineedit():
    le = QLineEdit()
    le.adjustSize()

    return le


class WR_GUI_frame(QWidget):
    def __init__(self, vocabulary):
        super().__init__()
        self.learn_flag = False
        self.review_flag = False
        self.current_line = 3
        self.setFixedSize(500, 350)
        self.setWindowTitle("Words Runner")
        #PC
        self.setWindowIcon(QIcon('./icon/title.ico'))
        self.read_vob(vocabulary)
        self.create_page()


    def read_vob(self, vocabulary):
        file_exl = openpyxl.load_workbook(vocabulary)
        sheet1_name = file_exl.get_sheet_names()[0]
        self.vob_sheet = file_exl.get_sheet_by_name(sheet1_name)
        '''
        Note:
        1. For toefl_vocabulary.xlsx, the words and their paraphrase are located
           from cell(2, 2) to (3647, 2).
        2. Words are located in column 2, and chinese meanings are in column 3.
        '''
        self.current_vob = openpyxl.Workbook()
        self.current_vob_sheet = self.current_vob.create_sheet(index = 0)
        print(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
        self.save_file_name = "vob_{}.xlsx".format(time.strftime("%Y-%m-%d-%H%M%S"))
        #self.current_vob.save(self.save_file_name)
        self.current_rvob_line = 0
        self.max_row = self.vob_sheet.max_row
        print("Words number: ", self.max_row)


    def create_page(self):
        self.learned_number = 0
        self.wrong_list = []
        self.current_word = self.vob_sheet.cell(self.current_line, 2).value
        self.current_para = self.vob_sheet.cell(self.current_line, 3).value

        self.word_chinese = create_label(self.current_para)
        self.word_chinese.setFixedSize(500, 25)
        self.word_qle = create_lineedit()
        self.word_qle.returnPressed.connect(lambda: self.ok_function(TYPE_REVIEW))

        self.numb_lab = create_label("Start from: ")
        self.numb_qle = create_lineedit()
        self.numb_jump = create_button("Go!", self.jump_function, None)

        self.current_numb_lab = create_label("Learned: ")
        self.current_numb_lab.setFixedSize(300, 25)

        self.learn_btn = create_button("Learned",
            lambda: self.ok_function(TYPE_LEARN),
            None)
        #self.learn_btn.setFixedSize(100, 25)
        self.review_btn = create_button("OK",
            lambda: self.ok_function(TYPE_REVIEW),
            None)
        self.prompt_btn = create_button("Prompt?", self.prompt_function, None)
        self.pass_btn = create_button("Pass", lambda: self.pass_function(cl = NO_JUMP), None)
        self.choose_vob_btn = create_button("Choose Vocabulary", self.choose_vocabulary, None)

        grid = QGridLayout()
        grid.setSpacing(10)
        grid.setContentsMargins(35, 80, 35, 80)
        self.setLayout(grid)

        grid.addWidget(self.numb_lab, 0, 3, 1, 1)
        grid.addWidget(self.numb_qle, 0, 4, 1, 1)
        grid.addWidget(self.numb_jump, 0, 5, 1, 1)

        grid.addWidget(self.word_qle, 1, 0, 1, 6)
        grid.addWidget(self.word_chinese, 2, 0, 1, 6)
        grid.addWidget(self.pass_btn, 4, 5)
        grid.addWidget(self.prompt_btn, 3, 0)
        grid.addWidget(self.learn_btn, 3, 4)
        grid.addWidget(self.review_btn, 3, 5)
        grid.addWidget(self.choose_vob_btn, 4, 0, 1, 2)

        grid.addWidget(self.current_numb_lab, 5, 0)


    def choose_vocabulary(self):
        file_name,  _ = QFileDialog.getOpenFileName(self, 'Choose Vocabulary', './', 'Excel files(*.xlsx , *.xls)')
        print("file: ", file_name)
        if file_name:
            self.read_vob(file_name)
            self.current_line = 1
            self.learned_number = 0
            self.wrong_list = []
            self.current_word = self.vob_sheet.cell(self.current_line, 2).value
            self.current_para = self.vob_sheet.cell(self.current_line, 3).value
            self.word_chinese.setText(self.current_para)
            self.word_qle.clear()


    def jump_function(self):
        number = self.numb_qle.text()
        if number.strip() != '':
            self.pass_function(cl = int(number))


    def ok_function(self, learn_type):
        type_words = self.word_qle.text().strip()
        #Learn mode
        if not self.review_flag:
            if type_words == self.current_word:
                if learn_type == TYPE_LEARN:
                    self.learn_flag = True
                    self.current_rvob_line += 1
                    print("learn: ", self.current_rvob_line)
                    self.current_vob_sheet.cell(self.current_rvob_line, 2).value = self.current_word
                    self.current_vob_sheet.cell(self.current_rvob_line, 3).value = self.current_para
                self.pass_function(cl = NO_JUMP)
                self.learned_number += 1
                self.current_numb_lab.setText("Learned: " + str(self.learned_number))
            else:
                warning = QMessageBox.warning(self, 'Ops!', 'Your anwser is wrong!' + self.current_word)
                if not self.current_line in self.wrong_list:
                    self.wrong_list.append(self.current_line)
                    print("wrong list: ", len(self.wrong_list), "line: ", self.current_line)
        #Review mode
        else:
            if type_words == self.current_word:
                self.review_function()
            else:
                warning = QMessageBox.warning(self, 'Ops!', 'Your anwser is wrong!' + self.current_word)
                if not self.current_line in self.wrong_list:
                    self.wrong_list.append(self.current_line)
                    print("wrong list: ", len(self.wrong_list), "line: ", self.current_line)


    def prompt_function(self):
        self.word_qle.setText(self.current_word)


    def pass_function(self, cl = NO_JUMP):
        if cl == NO_JUMP:
            self.current_line += 1
        else:
            self.current_line = cl

        print("current_line: ", self.current_line)
        if self.current_line <= self.max_row:
            self.current_word = self.vob_sheet.cell(self.current_line, 2).value
            self.current_para = self.vob_sheet.cell(self.current_line, 3).value
            #print("para: ", self.current_para)
            self.word_chinese.setText(self.current_para)
            self.word_qle.clear()
            QApplication.processEvents()
        else:
            if len(self.wrong_list) > 0:
                notice = QMessageBox.information(self, 'Review', 'Review wrong words!')
                self.review_function()
            else:
                notice = QMessageBox.information(self, 'Congratulations!', 'Congratulations! Words review Accomplished!')


    def review_function(self):
        self.review_flag = True
        if len(self.wrong_list) > 0:
            print(self.wrong_list[0])
            self.current_line = self.wrong_list[0]
            self.current_word = self.vob_sheet.cell(self.current_line, 2).value
            self.current_para = self.vob_sheet.cell(self.current_line, 3).value
            self.word_chinese.setText(self.current_para)
            self.word_qle.clear()
            self.wrong_list.pop(0)
        else:
            notice = QMessageBox.information(self, 'Congratulations!', 'Congratulations! Words review Accomplished!')


    def closeEvent(self, event):
        if self.learn_flag is True:
            self.current_vob.save(self.save_file_name)
        print("close")
        super().closeEvent(event)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    #Mac
    app.setWindowIcon(QIcon('./icon/title.ico'))

    ui_frame = WR_GUI_frame(VOCABULARY_PATH)
    ui_frame.show()

    sys.exit(app.exec_())
